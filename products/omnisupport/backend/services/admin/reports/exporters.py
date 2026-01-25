"""Report exporters for different formats."""

import io
import logging
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any
from uuid import UUID

logger = logging.getLogger(__name__)


class BaseExporter(ABC):
    """Base class for report exporters."""

    @abstractmethod
    async def export(self, report_data: dict, report_name: str) -> tuple[bytes, str]:
        """
        Export report data to specific format.

        Returns:
            Tuple of (file bytes, content type)
        """
        pass


class PDFExporter(BaseExporter):
    """Export reports to PDF format."""

    async def export(self, report_data: dict, report_name: str) -> tuple[bytes, str]:
        """Export report to PDF."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch, cm
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
            )
            from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        except ImportError:
            logger.error("reportlab not installed, using placeholder PDF")
            return self._generate_placeholder_pdf(report_data, report_name)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
        )
        heading_style = ParagraphStyle(
            'SectionHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=20,
            spaceAfter=10,
        )
        normal_style = styles['Normal']

        elements = []

        # Title
        elements.append(Paragraph(report_name, title_style))

        # Period
        period = report_data.get("period", {})
        if period:
            period_text = f"Период: {period.get('from', '')} - {period.get('to', '')}"
            elements.append(Paragraph(period_text, normal_style))
            elements.append(Spacer(1, 20))

        # Summary section
        if "totals" in report_data:
            elements.append(Paragraph("Сводка", heading_style))
            summary_data = self._format_summary_table(report_data["totals"])
            elements.append(self._create_table(summary_data))
            elements.append(Spacer(1, 20))

        if "summary" in report_data:
            elements.append(Paragraph("Сводка", heading_style))
            summary_data = self._format_summary_table(report_data["summary"])
            elements.append(self._create_table(summary_data))
            elements.append(Spacer(1, 20))

        if "team_summary" in report_data:
            elements.append(Paragraph("Командные показатели", heading_style))
            summary_data = self._format_summary_table(report_data["team_summary"])
            elements.append(self._create_table(summary_data))
            elements.append(Spacer(1, 20))

        # Operators table
        if "operators" in report_data:
            elements.append(Paragraph("Операторы", heading_style))
            operators_table = self._format_operators_table(report_data["operators"])
            elements.append(self._create_table(operators_table))
            elements.append(Spacer(1, 20))

        # Channels table
        if "channels" in report_data:
            elements.append(Paragraph("Каналы", heading_style))
            channels_table = self._format_channels_table(report_data["channels"])
            elements.append(self._create_table(channels_table))
            elements.append(Spacer(1, 20))

        # Distribution
        if "distribution" in report_data:
            elements.append(Paragraph("Распределение оценок", heading_style))
            dist_data = [["Оценка", "Количество"]]
            for score, count in sorted(report_data["distribution"].items()):
                dist_data.append([score, str(count)])
            elements.append(self._create_table(dist_data))
            elements.append(Spacer(1, 20))

        # Daily breakdown
        if "daily_breakdown" in report_data:
            elements.append(PageBreak())
            elements.append(Paragraph("Ежедневная статистика", heading_style))
            daily_table = self._format_daily_table(report_data["daily_breakdown"])
            elements.append(self._create_table(daily_table))

        if "daily_trend" in report_data:
            elements.append(PageBreak())
            elements.append(Paragraph("Динамика по дням", heading_style))
            daily_table = self._format_daily_trend_table(report_data["daily_trend"])
            elements.append(self._create_table(daily_table))

        # Footer
        elements.append(Spacer(1, 30))
        generated_at = report_data.get("generated_at", datetime.now().isoformat())
        elements.append(Paragraph(
            f"Отчёт сгенерирован: {generated_at}",
            ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.gray)
        ))

        doc.build(elements)
        buffer.seek(0)
        return buffer.read(), "application/pdf"

    def _generate_placeholder_pdf(self, report_data: dict, report_name: str) -> tuple[bytes, str]:
        """Generate a simple text-based placeholder when reportlab is not available."""
        content = f"Report: {report_name}\n\n"
        content += f"Data: {str(report_data)[:1000]}...\n"
        return content.encode('utf-8'), "text/plain"

    def _format_summary_table(self, data: dict) -> list[list[str]]:
        """Format summary data as table."""
        labels = {
            "conversations_total": "Всего диалогов",
            "conversations_new": "Новых диалогов",
            "conversations_resolved": "Решённых диалогов",
            "conversations_closed": "Закрытых диалогов",
            "messages_total": "Всего сообщений",
            "messages_inbound": "Входящих сообщений",
            "messages_outbound": "Исходящих сообщений",
            "customers_new": "Новых клиентов",
            "customers_active": "Активных клиентов",
            "avg_first_response_time": "Среднее время ответа (сек)",
            "avg_resolution_time": "Среднее время решения (сек)",
            "csat_score_avg": "Средний CSAT",
            "csat_responses": "Ответов CSAT",
            "total": "Всего",
            "resolved": "Решено",
            "total_responses": "Всего ответов",
            "average_score": "Средняя оценка",
            "nps": "NPS",
            "total_suggestions": "AI подсказок",
            "accepted": "Принято",
            "modified": "Изменено",
            "acceptance_rate": "Процент принятия",
            "total_conversations": "Всего диалогов",
            "total_resolved": "Всего решено",
        }

        table_data = [["Метрика", "Значение"]]
        for key, value in data.items():
            if key in labels and value is not None:
                if isinstance(value, float):
                    value = f"{value:.2f}"
                table_data.append([labels[key], str(value)])

        return table_data

    def _format_operators_table(self, operators: list[dict]) -> list[list[str]]:
        """Format operators data as table."""
        table_data = [["Оператор", "Диалоги", "Решено", "% решения", "Ср. ответ (с)"]]
        for op in operators:
            table_data.append([
                op.get("name", "Unknown"),
                str(op.get("conversations_total", 0)),
                str(op.get("conversations_resolved", 0)),
                f"{op.get('resolution_rate', 0)}%",
                str(op.get("avg_first_response_time", "-")),
            ])
        return table_data

    def _format_channels_table(self, channels: list[dict]) -> list[list[str]]:
        """Format channels data as table."""
        table_data = [["Канал", "Диалоги", "%", "Сообщения"]]
        for ch in channels:
            table_data.append([
                ch.get("channel", "Unknown"),
                str(ch.get("conversations", 0)),
                f"{ch.get('percentage', 0)}%",
                str(ch.get("messages", 0)),
            ])
        return table_data

    def _format_daily_table(self, daily: list[dict]) -> list[list[str]]:
        """Format daily breakdown as table."""
        table_data = [["Дата", "Новых", "Решено", "Сообщений"]]
        for day in daily:
            table_data.append([
                day.get("date", ""),
                str(day.get("conversations_new", 0)),
                str(day.get("conversations_resolved", 0)),
                str(day.get("messages_total", 0)),
            ])
        return table_data

    def _format_daily_trend_table(self, daily: list[dict]) -> list[list[str]]:
        """Format daily trend as table."""
        if not daily:
            return [["Нет данных"]]

        # Determine columns from first item
        first = daily[0]
        columns = list(first.keys())
        headers = {
            "date": "Дата",
            "count": "Количество",
            "score": "Оценка",
            "responses": "Ответов",
            "suggestions": "Подсказок",
            "accepted": "Принято",
            "modified": "Изменено",
        }

        table_data = [[headers.get(c, c) for c in columns]]
        for day in daily:
            row = [str(day.get(c, "")) for c in columns]
            table_data.append(row)
        return table_data

    def _create_table(self, data: list[list[str]]) -> Table:
        """Create styled table."""
        from reportlab.lib import colors
        from reportlab.platypus import Table, TableStyle

        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        return table


class ExcelExporter(BaseExporter):
    """Export reports to Excel format."""

    async def export(self, report_data: dict, report_name: str) -> tuple[bytes, str]:
        """Export report to Excel."""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl not installed, falling back to CSV")
            exporter = CSVExporter()
            return await exporter.export(report_data, report_name)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчёт"

        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=16)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        current_row = 1

        # Title
        ws.cell(row=current_row, column=1, value=report_name).font = title_font
        current_row += 2

        # Period
        period = report_data.get("period", {})
        if period:
            ws.cell(row=current_row, column=1, value=f"Период: {period.get('from', '')} - {period.get('to', '')}")
            current_row += 2

        # Summary
        summary_data = report_data.get("totals") or report_data.get("summary") or report_data.get("team_summary")
        if summary_data:
            ws.cell(row=current_row, column=1, value="Сводка").font = header_font
            current_row += 1
            current_row = self._write_dict_as_rows(ws, summary_data, current_row, header_fill, header_font_white, thin_border)
            current_row += 1

        # Operators
        if "operators" in report_data:
            ws.cell(row=current_row, column=1, value="Операторы").font = header_font
            current_row += 1
            current_row = self._write_list_as_table(ws, report_data["operators"], current_row, header_fill, header_font_white, thin_border)
            current_row += 1

        # Channels
        if "channels" in report_data:
            ws.cell(row=current_row, column=1, value="Каналы").font = header_font
            current_row += 1
            current_row = self._write_list_as_table(ws, report_data["channels"], current_row, header_fill, header_font_white, thin_border)
            current_row += 1

        # Distribution
        if "distribution" in report_data:
            ws.cell(row=current_row, column=1, value="Распределение").font = header_font
            current_row += 1
            current_row = self._write_dict_as_rows(ws, report_data["distribution"], current_row, header_fill, header_font_white, thin_border)
            current_row += 1

        # Daily data in separate sheet
        daily_data = report_data.get("daily_breakdown") or report_data.get("daily_trend")
        if daily_data:
            daily_ws = wb.create_sheet(title="Ежедневно")
            self._write_list_as_table(daily_ws, daily_data, 1, header_fill, header_font_white, thin_border)

        # Adjust column widths
        for ws_item in wb.worksheets:
            for column_cells in ws_item.columns:
                length = max(len(str(cell.value or "")) for cell in column_cells)
                ws_item.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    def _write_dict_as_rows(self, ws, data: dict, start_row: int, header_fill, header_font, border) -> int:
        """Write dictionary as rows."""
        labels = {
            "conversations_total": "Всего диалогов",
            "conversations_new": "Новых диалогов",
            "conversations_resolved": "Решённых",
            "avg_first_response_time": "Среднее время ответа (сек)",
            "avg_resolution_time": "Среднее время решения (сек)",
            "csat_score_avg": "Средний CSAT",
            "total_suggestions": "AI подсказок",
            "acceptance_rate": "Процент принятия",
        }

        # Headers
        ws.cell(row=start_row, column=1, value="Метрика").fill = header_fill
        ws.cell(row=start_row, column=1).font = header_font
        ws.cell(row=start_row, column=2, value="Значение").fill = header_fill
        ws.cell(row=start_row, column=2).font = header_font
        start_row += 1

        for key, value in data.items():
            if value is not None:
                ws.cell(row=start_row, column=1, value=labels.get(key, key)).border = border
                ws.cell(row=start_row, column=2, value=value).border = border
                start_row += 1

        return start_row

    def _write_list_as_table(self, ws, data: list[dict], start_row: int, header_fill, header_font, border) -> int:
        """Write list of dicts as table."""
        if not data:
            return start_row

        headers = list(data[0].keys())
        header_labels = {
            "name": "Имя",
            "email": "Email",
            "conversations_total": "Диалоги",
            "conversations_resolved": "Решено",
            "resolution_rate": "% решения",
            "avg_first_response_time": "Ср. ответ (с)",
            "channel": "Канал",
            "conversations": "Диалоги",
            "percentage": "%",
            "messages": "Сообщения",
            "date": "Дата",
            "count": "Количество",
            "score": "Оценка",
        }

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col, value=header_labels.get(header, header))
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        start_row += 1

        # Write data
        for row_data in data:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=start_row, column=col, value=row_data.get(header, ""))
                cell.border = border
            start_row += 1

        return start_row


class CSVExporter(BaseExporter):
    """Export reports to CSV format."""

    async def export(self, report_data: dict, report_name: str) -> tuple[bytes, str]:
        """Export report to CSV."""
        import csv

        buffer = io.StringIO()
        writer = csv.writer(buffer)

        # Write metadata
        writer.writerow(["Report", report_name])
        period = report_data.get("period", {})
        if period:
            writer.writerow(["Period", f"{period.get('from', '')} to {period.get('to', '')}"])
        writer.writerow([])

        # Write summary
        summary = report_data.get("totals") or report_data.get("summary") or report_data.get("team_summary")
        if summary:
            writer.writerow(["Summary"])
            writer.writerow(["Metric", "Value"])
            for key, value in summary.items():
                if value is not None:
                    writer.writerow([key, value])
            writer.writerow([])

        # Write operators
        if "operators" in report_data:
            writer.writerow(["Operators"])
            operators = report_data["operators"]
            if operators:
                headers = list(operators[0].keys())
                writer.writerow(headers)
                for op in operators:
                    writer.writerow([op.get(h, "") for h in headers])
            writer.writerow([])

        # Write channels
        if "channels" in report_data:
            writer.writerow(["Channels"])
            channels = report_data["channels"]
            if channels:
                headers = list(channels[0].keys())
                writer.writerow(headers)
                for ch in channels:
                    writer.writerow([ch.get(h, "") for h in headers])
            writer.writerow([])

        # Write daily data
        daily = report_data.get("daily_breakdown") or report_data.get("daily_trend")
        if daily:
            writer.writerow(["Daily Data"])
            if daily:
                headers = list(daily[0].keys())
                writer.writerow(headers)
                for day in daily:
                    writer.writerow([day.get(h, "") for h in headers])

        content = buffer.getvalue()
        return content.encode('utf-8'), "text/csv"


def get_exporter(format: str) -> BaseExporter:
    """Get exporter for format."""
    exporters = {
        "pdf": PDFExporter,
        "excel": ExcelExporter,
        "csv": CSVExporter,
    }
    exporter_class = exporters.get(format.lower(), CSVExporter)
    return exporter_class()
